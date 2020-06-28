import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, colors, Alignment, Font
from datetime import datetime
import alarm


def set_border(top, bottom, left, right):
    border = Border(top=Side(border_style=top, color=colors.BLACK),
                    bottom=Side(border_style=bottom, color=colors.BLACK),
                    left=Side(border_style=left, color=colors.BLACK),
                    right=Side(border_style=right, color=colors.BLACK))
    return border


class DataXl:
    task_list = []  # 存放任务

    def __init__(self):
        self.wb = Workbook()
        self.wb = openpyxl.load_workbook("C:/Users/FYH/PycharmProjects/记事本/notebook.xlsx")
        self.ws = self.wb.active
        self.ws.title = "To Be Discipline"
        self.ws.sheet_properties.tabColor = "50616d"
        for col in range(ord('B'), ord('I')):
            self.ws.column_dimensions[chr(col)].width = 20.0
        for row in range(2, 8):
            self.ws.row_dimensions[row].height = 80.0
        for col in range(2, 9):
            for row in range(2, 8):
                self.ws.cell(row, col).border = set_border("thick", 'thick', 'thick', 'thick')
        for col in range(2, 9):
            self.ws.cell(2, col).value = f"星期{col - 1}"
            self.ws.cell(2, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.ws.cell(2, col).font = Font(size=13.9)

    def save_task(self, today_max_row):  # 传入任务数
        dayOfWeek = datetime.now().isoweekday()  # 今天星期几，用于确定列
        for row in range(3 + today_max_row - 1, 8):  # 清空上周今天的任务
            self.ws.cell(row, dayOfWeek + 1).value = ""
        for row in range(3, 3 + today_max_row):  # 从第一行开始写任务数个数据
            self.ws.cell(row, dayOfWeek + 1).alignment = Alignment(horizontal='center', vertical='center',
                                                                   wrap_text=True)
            task = self.task_list[row - 3][0]  # my_task  # input(f"task{row - 2}: ")  # 输入任务
            start_time = self.task_list[row - 3][2]  # my_start  # input("startTime: ")  # 输入开始时间xx:xx
            end_time = self.task_list[row - 3][1]  # my_end  # input("endTime: ")  # 输入结束时间xx:xx
            end_hour, end_minute = end_time.split(':')  # 获取结束的时、分
            end_time_str = end_time.split(':')[0] + end_time.split(':')[1]  # 获得string类型结束时间
            self.ws.cell(row, dayOfWeek + 1).value = task  # 写入task
            self.ws.cell(row, dayOfWeek + 1).value += '\n' + start_time + '-' + end_time  # 写入开始、结束时间
            self.task_list.append([task, end_time])  # 在task_list里依次尾插任务
        self.wb.save("Notebook.xlsx")  # 存储本次更改

    def times_up(self, today_max_row):
        for row in range(3, 3 + today_max_row):
            alarm.time_it_is(self.task_list[row - 3][1].split(':')[0],
                             self.task_list[row - 3][1].split(':')[1])  # 传入时和分

    def execute_it(self, today_max_row):
        # self.save_task(today_max_row)
        self.times_up(today_max_row)


# test = DataXl()
# test.execute_it(2)
